"""
    匹配两页的开挖报告
    
"""
from multiprocessing.dummy import Value
import re
from tabnanny import check
from numpy import s_, str_
import openpyxl
from openpyxl.workbook import Workbook
import datetime
import win32com.client as win32
import os
import traceback
import math,random
from src.mypackage import r_generator as rg
from src.mypackage.LOG_DATA  import LOG_DICT,RISKY_EVA_C,RISKY_EVA_S
from src.mypackage import interraction_terminal


"""=========================编辑生成全部用于替换的列表索引文件replacements======================"""

def expand_all_tables(workbook:Workbook, doc, report_name:str,rm_dict)->dict[str,list[int]]:
    """按照读取到的分项记录数量，复制报告页张数。返回宏观、穿跨越的多页组织数量列表"""
    f_dict:dict[str,list[int]]={}
    #   资料审查
    sheet=workbook['管段清单']
    log_dict:dict=rg.get_col_in_sheet(sheet)
    rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    times:int =len(rows)
    f_dict['资料审查']=[1]*times
    if times>1:
        rg.copy_and_insert_report_bookmark(doc , '资料审查记录', times)
    rg.replace_text(doc, '复制资料审查','',2)  
    
    #   宏观检查
    # sheet = workbook['宏观检查记录']
    # log_dict:dict = rg.get_col_in_sheet(sheet)
    # times:int = len(rg.get_rows_in_sheet(report_name, sheet ,log_dict['报告编号'])) 
    
    # rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    times = len(rm_dict['宏观人员'])
    f_dict['宏观检查']=[1]*times
    # f_dict['宏观检查']=[]
    # for row in rows:
    #     son_rows = rg.get_rows_in_sheet(sheet[log_dict['记录自编号']+row].value,sheet,log_dict['所属记录编号'])
    #     count = 0
    #     for son_row in son_rows:
    #         v:str =sheet[log_dict['检查项目类别']+son_row].value
    #         count += len(v.split(', '))
    #     time = math.ceil(count/14)
    #     times+=time
    #     f_dict['宏观检查'].append(time)
    if times>1:
        rg.copy_and_insert_report_bookmark(doc , '宏观检查记录', times)
    # rg.replace_text(doc, '复制宏观检查','',2)

    #   开挖检测
    sheet = workbook['开挖检测记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times = len(rm_dict['开挖人员'])
    f_dict['开挖检测']=[1]*times
    if times>1:
        rg.copy_and_insert_report_bookmark(doc , '开挖直接检验记录', times,1)   # 这里复制两页
    # rg.replace_text(doc, '复制开挖检测','',2)  
       
    #   穿跨越检查
    times = len(rm_dict['穿、跨越人员'])
    f_dict['穿跨越检查']=[1]*times
    if times>1:
        rg.copy_and_insert_report_bookmark(doc , '穿、跨越检查记录', times)
    # rg.replace_text(doc, '复制穿跨越检查','',2)  
    return f_dict
    
    #   整理删除页面
    # rg.delete_page_by_text(doc, '待删除')

"""
========================执行替换=======================
"""
#   替换文本
def do_replace(doc , replacements1:list[tuple[str,str]],replacements2:list[tuple[str,str]]=[])->None:
    """替换所有文本，先替换全局，再替换单次"""
    for target_text, replacement_text in replacements2:
        rg.replace_text(doc, target_text, replacement_text,2 )
    for target_text, replacement_text in replacements1:
        rg.replace_text(doc, target_text, replacement_text )

#   签字函数
def make_sign_dig_log(workbook:Workbook,doc,report_name:str,path:str,rm_dict:dict[str,list[list[str]]],times_dict,gd_dict)->dict:
    """编制签名图片、开挖图片、路由图替换的索引，依照输入的分项扩张字典来生成签字索引"""
    # 签资料审查：遍历所有的宏观和开挖记录检验人员，通过集合无序化，资料最多签2人
    sign_dict:dict = {}
    all_names:set[str] = {name_str for t_list in rm_dict['宏观人员'] for name_str in t_list}
    # sheet = workbook['宏观检查记录']
    # log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
    # rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # for row in rows:
    #     if sheet[log_dict['检验人员']+row].value:
    #         names:list[str] =sheet[log_dict['检验人员']+row].value.split(', ')
    #     for name in names:
    #         all_names.add(name)

    # sheet = workbook['开挖检测记录']
    # log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
    # rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # for row in rows:
    #     if sheet[log_dict['检验人员']+row].value:
    #         names:list[str] =sheet[log_dict['检验人员']+row].value.split(',')
    #     else:
    #         names=[]
    #     for name in names:
    #         all_names.add(name)
    
    sign_dict['签字'] = []  
    all_names_list=list(all_names)
    sign_dict['签字'] += all_names_list[0:1]*len(times_dict['资料审查'])
    # sign_dict['签字'] += all_names_list[0:2]

    # 签宏观检查报告
    # sheet =workbook['宏观检查记录']
    # log_dict = rg.get_col_in_sheet(sheet)
    # rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # for lst,row in enumerate(rows):
    #     names_set:set[str] = set()
    #     if  sheet[log_dict['检验人员']+row].value:
    #         for name in sheet[log_dict['检验人员']+row].value.split(', '):
    #             names_set.add(name)
    #     lenth = len(names_set)
    #     temp_list = list(names_set)
    #     for _ in range(times_dict['宏观检查'][lst]):
    #         sign_dict['签字'] += temp_list[:2]  # 宏观检查
    #         if lenth<2:
    #             sign_dict['签字'] += ['空白']*(2-lenth)
    for names in rm_dict['宏观人员']:
        sign_dict['签字'] += names[:2] 
        if len(names)<2:
            sign_dict['签字'] += ['空白']*(2-len(names))
    
    #   签开挖检验报告
    sheet =workbook['开挖检测记录']
    log_dict = rg.get_col_in_sheet(sheet)
    # rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    rows=[]
    for gd_no in gd_dict.keys():
        temp_list:list[str] = []
        temp_list = rg.get_rows_in_sheet(gd_no,sheet,log_dict['管道编码'])
        if len(temp_list):
            rows+=temp_list
    sign_dict['开挖检验照片']=[] 
    for row in rows:
        sign_dict['开挖检验照片'].append(sheet[log_dict['记录自编号']+row].value) 
    #     names_set = set()
    #     if sheet[log_dict['检验人员']+row].value:
    #         for name in sheet[log_dict['检验人员']+row].value.split(','):
    #             names_set.add(name)
    #     lenth = len(names_set)
    #     temp_list = list(names_set)
    #     sign_dict['签字'] += temp_list[:2]  
    #     if lenth<2:
    #         sign_dict['签字'] += ['空白']*(2-lenth)
    #     sign_dict['签字'] += temp_list[:2]  
    #     if lenth<2:
    #         sign_dict['签字'] += ['空白']*(2-lenth)
    for names in rm_dict['开挖人员']:
        for _ in range(2): 
            sign_dict['签字'] += names[:2]
            if len(names)<2:
                sign_dict['签字'] += ['空白']*(2-len(names))


    #   穿跨越记录
    # sheet =workbook['宏观检查记录']
    # log_dict = rg.get_col_in_sheet(sheet)
    # rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # i:int = 0
    # if all(x==0 for x in times_dict['穿跨越检查']):
    #     sign_dict['签字']+=list(all_names)[:2]
    # else:
    #     for row in rows:
    #         names_set:set = set()
    #         if sheet[log_dict['检验人员']+row].value:
    #             for name in sheet[log_dict['检验人员']+row].value.split(', '):
    #                 names_set.add(name)
    #         for _ in range(times_dict['穿跨越检查'][i]):
    #             lenth = len(names_set)
    #             temp_list = list(names_set)
    #             sign_dict['签字'] += temp_list[:2]  
    #             if lenth<2:
    #                 sign_dict['签字'] += ['空白']*(2-lenth)
    #         i += 1
    for names in rm_dict['穿、跨越人员']:
        sign_dict['签字'] += names[:2] 
        if len(names)<2:
            sign_dict['签字'] += ['空白']*(2-len(names))
    
    #   风险评估记录及打分表
    any_name=list(all_names)[0]
    sign_dict['签字'] += [any_name]*16

    return sign_dict

#   宏观检查的行号梳理
def make_macro_index(workbook:Workbook,report_name:str,gd_dict:dict[str,str])->dict[str,list[str]]:
    """按照输入的管道编码，返回表头行号——子记录行号列表构成的字典"""
    macro_index:dict[str,list[str]]={}
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    for cell in sheet[log_dict['管道编码']]:
        c_v = str(cell.value) if cell.value else ''
        if c_v and (c_v in gd_dict.keys()):
            s_row =str(cell.row)
            parent_no =sheet[log_dict['所属记录编号']+s_row].value
            rows =rg.get_rows_in_sheet(parent_no,sheet,log_dict['记录自编号'])
            if rows[0] not in macro_index.keys():
                macro_index[rows[0]]=[]
            macro_index[rows[0]]+=[s_row]
    return macro_index

#   完成索引
def make_replacement_index(workbook:Workbook,report_name:str,gd_dict:dict[str,str])->dict:
    """全体替换内容，主要调整函数"""
    replacements:dict={}
    replacements['文本'] = []
    
    sheet=workbook['管道基本信息']
    log_dict= rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    global_name = sheet[log_dict['管道名称']+rows[0]].value
    
    #   资料审查记录
    replacements['资料审查记录']=[]
    sheet=workbook['管段清单']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])  
    for row in rows:
        
        temp_list:list[tuple]=[]
        for key in LOG_DICT['资料审查记录生成']:
            key=key.replace('+','')
            if key in log_dict.keys():
                if key in ['设计压力','操作压力']:
                    temp_list+=[(f'+{key}',f'{sheet[log_dict[key]+row].value}MPa')]
                elif '年限' in key:
                    temp_list+=[(f'+{key}',f'{sheet[log_dict[key]+row].value}年')]
                else:
                    temp_list+=[(f'+{key}',sheet[log_dict[key]+row].value)]
        # temp_list+=[    #   管网一批次
        #         ('+管段名称',sheet[log_dict['工程名称']+row].value),
        #         ('+安全管理资料审查','无专项应急预案及演练记录'),
        #         ('+运行状况资料审查','除巡查记录，抢修记录外，未见其他资料'),
        #         ('+资料审查问题记载','除以上问题外，未见管道历次年度检查资料，本次为首次全面检验'),
        #         # ('+检验日期','2024年07月09日'),
        #         ('+长度',f'{sheet[log_dict['长度（m）']+row].value}m'),
                
        #     ]  
        # if sheet[log_dict['工程名称']+row].value!='使用单位指定管段':
        #     temp_list+=[
        #         ('+技术档案资料审查','仅见竣工图，未见设计文件、质量证明文件、监督检验证书及安装、改造、修理资料')
        #     ]
        # else:
        #     temp_list+=[('+技术档案资料审查','无资料')]  
        temp_list+=[    #   新繁，大丰
                ('+管段名称',sheet[log_dict['工程名称']+row].value),
                # ('+运行状况资料审查','有日常运行维护记录，未见管道历次年度检查报告'),
                # ('+资料审查问题记载','除以上问题外，未见管道历次年度检查资料，本次为首次全面检验'),
                # ('+检验日期','2024年07月09日'),
                ('+长度',f'{sheet[log_dict['长度']+row].value}m'),
            ] 
        # lst_no=int(sheet[log_dict['序号']+row].value)
        # last_report_no=['GD202200194','GD202200195','GD202200196','GD202200197','GD202200198']
        # if lst_no in [67,114,694,695,697,698,699,700] or (lst_no>=679 and lst_no<691):  #  新繁
        # # if lst_no in [5,86,169]:  #  大丰
        #     temp_list+=[
        #         ('+安全管理资料审查','有安全管理规章制度与安全操作规则、该管道为首次检验，有监检报告'),
        #         ('+检验周期','本次为首次检验')
        #         ] 
        # else:
        #     temp_list+=[
        #         ('+安全管理资料审查','有安全管理规章制度与安全操作规则、有上次定期检验报告'),
        #         ('+检验周期','3年'),
        #         ('+上次检验日期','2022年6月'),
        #         # ('+上次报告编号','GD2023000009'),   #   大丰
        #         ('+上次报告编号',random.choice(last_report_no)),   #   新繁
        #         # ] 
        
        e_name=sheet[log_dict['工程名称']+row].value
        fdate=sheet[log_dict['竣工验收日期']+row].value
        # temp_text='无资料，仅有GIS系统位置信息' if '使用单位指定管段' in e_name else '仅见竣工图，未见设计文件、质量证明文件、监督检验证书及安装、改造、修理资料'
        # temp_list+=[('+技术档案资料审查',temp_text)]
        # temp_list+=[('+资料审查问题记载','，'.join([t for t in temp_text_list if '无' in t]+['无管道历次年度检查资料']))]
        replacements['资料审查记录'].append(temp_list)
    
    #   宏观检查记录
    replacements['宏观检查记录']=[]
    replacements['宏观人员']=[]
    temp_list:list[tuple]= []
    sheet =workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    # rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    marco_dict=make_macro_index(workbook,report_name,gd_dict)
    for row,son_rows in marco_dict.items():
        temp_list=[]            
        no_set:set[str] = set()     #   管道编号的集合
        # record_num = sheet[log_dict['记录自编号']+row].value
        # son_rows = rg.get_rows_in_sheet(record_num,sheet,log_dict['所属记录编号'])  # 获取子记录行索引
        temp_count=0
        for son_row in son_rows:
            key_str:str = sheet[log_dict['检查项目类别']+son_row].value
            no_set.add(str(sheet[log_dict['管道编码']+son_row].value))
            keys = key_str.split(', ')
            for key in keys:
                temp_count+=1
                if '穿' in key:
                    checked:str = f"{sheet[log_dict['穿跨越类型']+son_row].value}{sheet[log_dict['穿跨越长度']+son_row].value}米"
                elif '阀门井' in key and sheet[log_dict['管道埋深']+son_row].value is not None:
                    if sheet[log_dict['阀门井']+son_row].value is not None:
                        checked:str = f"{sheet[log_dict['阀门井']+son_row].value},埋深{sheet[log_dict['管道埋深']+son_row].value}米"
                    else:
                        checked:str = f"埋深{sheet[log_dict['管道埋深']+son_row].value}米"
                elif '位置与走向' in key and sheet[log_dict['管道埋深']+son_row].value is not None:
                    if sheet[log_dict['位置与走向']+son_row].value is not None:
                        checked:str = f"{sheet[log_dict['位置与走向']+son_row].value},埋深{sheet[log_dict['管道埋深']+son_row].value}米"
                    else:
                        checked:str = f"埋深{sheet[log_dict['管道埋深']+son_row].value}米"
                elif sheet[log_dict[key]+son_row].value is not None:
                    checked:str = f'{sheet[log_dict[key]+son_row].value}'
                else:
                    checked:str = '/'

                temp_list += [
                    (f"&号{temp_count%14}",temp_count),
                    (f"&项目类别{temp_count%14}",key),
                    (f"&坐标{temp_count%14}",f"{sheet[log_dict['坐标X']+son_row].value},{sheet[log_dict['坐标Y']+son_row].value}"),
                    (f"&地标、位置{temp_count%14}",sheet[log_dict['地表参照及位置描述']+son_row].value),
                    (f"&检查结果{temp_count%14}",checked),

                ]
        temp_head_list = [  # 表头部分的固定内容
                    # ('+管道名称','、'.join([gd_dict[no] for no in no_set])),
                    # ('+管道编号','、'.join(no_set)),
                    ('+管段','、'.join([gd_dict[no] for no in no_set][:2])),
                    ('+管道编号','、'.join([no for no in no_set][:2])),
                    ('+设备名称型号',sheet[log_dict['设备名称型号']+row].value),
                    ('+设备编号',sheet[log_dict['设备编号']+row].value),
                   
                    ('+环境条件',sheet[log_dict['环境条件']+row].value),
                    ]
        if CONFIG['是否检验签字']:
            temp_head_list+=[ ('+检验日期',sheet[log_dict['检验日期']+row].value),]
        any_count=math.ceil(len(temp_list)/5/14)
        for i in range(any_count):
            and_list = temp_list[70*i:70*(i+1)]
            lenth:int = int(len(and_list)/5)
            if  lenth < 14:
                n=(lenth+1)%14
                and_list+=[
                    (f"&号{n}",'/'),
                    (f"&项目类别{n}",'/'),
                    (f"&坐标{n}",'/'),
                    (f"&地标、位置{n}",'/'),
                    (f"&检查结果{n}",'/'),
                ]
            replacements['宏观检查记录'].append(temp_head_list+and_list) 
            replacements['宏观人员'].append(sheet[log_dict['检验人员']+row].value.split(', '))

    
    #   开挖检验报告
    replacements['开挖记录首页']=[]
    replacements['开挖检验记录']=[]
    replacements['开挖人员']=[]
    sheet = workbook['开挖检测记录']
    log_dict = rg.get_col_in_sheet(sheet)  #获取表头索引
    rows:list=[]
    for gd_no in gd_dict.keys():
        temp_list:list[str] = []
        temp_list = rg.get_rows_in_sheet(gd_no,sheet,log_dict['管道编码'])
        if len(temp_list):
            rows+=temp_list
    # rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    for row in rows:
        temp_list_0:list[tuple] = []
        temp_list:list[tuple] = []
       
        temp_list += rg.make_change_text_for_heading(sheet,row,'开挖检测记录',log_dict)
        gd_no:str=str(sheet[log_dict['管道编码']+row].value)
        temp_list += [
            ('+管道名称',sheet[log_dict['管道名称']+row].value),
            ('+管道编号',gd_no),
            ('+探坑编号',sheet[log_dict['探坑编号']+row].value),
            ]
        temp_list += rg.make_change_text_for_option(sheet,row,'开挖检测记录',log_dict)
        temp_list_0 += [    #  开挖首页的表头
            ('+管道名称',gd_dict[gd_no]),
            # ('+管道编号',gd_no),
            ('+管道规格',sheet[log_dict['管道规格']+row].value),
            ('+实际检验日期',sheet[log_dict['检验日期']+row].value),
            ('+探坑编号',sheet[log_dict['探坑编号']+row].value),
            ('+探坑位置',sheet[log_dict['探坑位置']+row].value),
            ('+探坑规格',sheet[log_dict['探坑规格（m）']+row].value),
            ('+地表状况',sheet[log_dict['地形、地貌、地物描述']+row].value),
            ('+检验情况',f"检验情况：{sheet[log_dict['缺陷描述']+row].value if sheet[log_dict['缺陷描述']+row].value else '管体无缺陷'}{('，'+str(sheet[log_dict['备注']+row].value)) if sheet[log_dict['备注']+row].value else ''}。"),
            ('+检验结论',f"检验结论：根据GB/T 43922-2024《在役聚乙烯燃气管道检验与评价》安全状况等级评定为{sheet[log_dict['结论']+row].value}"),
            ('+检验日期',sheet[log_dict['检验日期']+row].value),
            ]
        if CONFIG['是否检验签字']:
            temp_list+=[('+检验日期',sheet[log_dict['检验日期']+row].value)]
        v1 = sheet[log_dict['备注']+row].value
        if v1 is None:
            v1 = ''
        else:
            v1 = f"{rg.check_text(v1)}。"
        v2 = rg.check_text(sheet[log_dict['结论']+row].value)
        temp_list += [('+备注',f"备注：{v1 if v1 else '/'}"),]
        replacements['开挖检验记录'].append(temp_list)
        replacements['开挖记录首页'].append(temp_list_0)
        replacements['开挖人员'].append(sheet[log_dict['检验人员']+row].value.split(','))
    
    #   穿、跨越检查
    replacements['穿、跨越检查记录'] = []
    replacements['穿、跨越人员'] = []
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    # rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    temp_count = 0  # 报告标题的编号
    report_no:int = 0       # 穿跨越报告的序号
                
    for row,son_rows in marco_dict.items():
        # 本地确认
        exists_set= set(sheet[log_dict['穿跨越类型']+son_row].value for son_row in son_rows if sheet[log_dict['穿跨越类型']+son_row].value is not None )
        if not exists_set:
            continue
        else:
            report_no+=1
            record_num = sheet[log_dict['记录自编号']+row].value
            # son_rows = rg.get_rows_in_sheet(record_num,sheet,log_dict['所属记录编号'])
            son_rows1:list[str]=[]
            son_rows2:list[str]=[]
            no_set:set[str] = set()     #   管道编号的集合
            for son_row in son_rows:
                v = sheet[log_dict['穿跨越类型']+son_row].value
                no_set.add(str(sheet[log_dict['管道编码']+son_row].value))
                if v == '跨越':
                    son_rows1.append(son_row)
                elif v=='穿越':
                    son_rows2.append(son_row)
            cap1:int=len(son_rows1)
            cap2:int=len(son_rows2)
            pages:int = max(math.ceil(cap1/8),math.ceil(cap2/10))
            cr_result1 = set(sheet[log_dict['穿、跨越河流']+son_row].value for son_row in son_rows if sheet[log_dict['穿、跨越河流']+son_row].value is not None)
            cr_result2 = set(sheet[log_dict['穿、跨越公路']+son_row].value for son_row in son_rows if sheet[log_dict['穿、跨越公路']+son_row].value is not None)
            b_result = '穿、跨越段仅宏观检验,未发现异常'
            if cr_result1 is not None and '保护设施完好' in cr_result1:
                b_result = '保护设施完好'
            if cr_result2 is not None and '保护设施完好' in cr_result2:
                b_result = '保护设施完好'
            page:int = 0        # 标志单一报告的页序数
            for _ in range(pages):
                page += 1
                
                temp_list:list[tuple]=[]
                temp_list += [      # 表头和结论
                        ('+管段','、'.join([gd_dict[no] for no in no_set][:2])),
                        # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                        ('+环境条件',sheet[log_dict['环境条件']+row].value),
                        ('+管道编号','、'.join([no for no in no_set][:2])),
                        # ('+检查结论',f"检查结论：{b_result}"),
                        ]
                if CONFIG['是否检验签字']:
                    temp_list+=[('+检验日期',sheet[log_dict['检验日期']+row].value)]
                # 跨越填表
                temp_count = 0
                if len(son_rows1) >0:
                    for son_row in son_rows1[:8]:
                        temp_count +=1
                        temp_list +=[
                            (f"&号{temp_count}",temp_count+8*(page-1)),
                            (f'&长度{temp_count}',sheet[log_dict['穿跨越长度']+son_row].value),
                            (f'&发现问题及位置描述{temp_count}',f'{sheet[log_dict['地表参照及位置描述']+son_row].value}（{sheet[log_dict['坐标X']+son_row].value}，{sheet[log_dict['坐标Y']+son_row].value}）' ),
                            (f'&备注{temp_count}','/')
                        ]
                    if temp_count < 8:
                        temp_count += 1
                        temp_list += [
                            (f"&号{temp_count}",'/'),
                            (f'&长度{temp_count}','/'),
                            (f'&发现问题及位置描述{temp_count}','/' ),
                            (f'&备注{temp_count}','/')
                        ]
                    son_rows1 = son_rows1[8:]
                else:
                    temp_list += [
                            ("&号1",'/'),
                            ('&长度1','/'),
                            ('&发现问题及位置描述1','/' ),
                            ('&备注1','/')
                        ]
                # 穿越填表
                temp_count = 0
                if len(son_rows2) >0:
                    for son_row in son_rows2[:10]:
                        temp_count +=1
                        temp_list +=[
                            (f"$号{temp_count}",temp_count+10*(page-1)),
                            (f'$长度{temp_count}',sheet[log_dict['穿跨越长度']+son_row].value),
                            (f'$发现问题及位置描述{temp_count}',f'{sheet[log_dict['地表参照及位置描述']+son_row].value}（{sheet[log_dict['坐标X']+son_row].value}，{sheet[log_dict['坐标Y']+son_row].value}）' ),
                            (f'$备注{temp_count}','/')
                        ]
                    if temp_count < 10:
                        temp_count += 1
                        temp_list += [
                            (f"$号{temp_count}",'/'),
                            (f'$长度{temp_count}','/'),
                            (f'$发现问题及位置描述{temp_count}','/' ),
                            (f'$备注{temp_count}','/')
                        ]
                        son_rows2 = son_rows2[10:]
                else:
                    temp_list += [
                            ("$号1",'/'),
                            ('$长度1','/'),
                            ('$发现问题及位置描述1','/' ),
                            ('$备注1','/')
                        ]
                replacements['穿、跨越检查记录'].append(temp_list)
                replacements['穿、跨越人员'].append(sheet[log_dict['检验人员']+row].value.split(', '))
    # 风险评估
    risk_point_s:dict[str,list[int]]={}
    risk_point_c:dict[str,list[int]]={}
    sheet =workbook['风险评估']
    log_dict=rg.get_col_in_sheet(sheet)
    rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    replacements['风险预评估']=[]
    risk_point_c['风险预评估']=[]
    risk_point_c['风险再评估']=[]
    risk_point_s['风险预评估']=[]
    risk_point_s['风险再评估']=[]
    row = rows[0]
    for key_str,list_dict in RISKY_EVA_S.items():
        risk_score:int=0 
        for son_key,son_tuple in list_dict.items():
            v= sheet[log_dict[son_key]+row].value # 表格实际内容
            for option,score in son_tuple:
                if isinstance(option,tuple): # 如果键是区间（元组）
                    if v>=option[0] and v<option[1]:
                        risk_score += score
                        risk_point_s['风险预评估'].append(score)
                else:
                    if v == option:
                        risk_score += score
                        risk_point_s['风险预评估'].append(score)
        replacements['风险预评估'].append(risk_score)
    for key_str,list_dict in RISKY_EVA_C.items():
        risk_score:int=0 
        v= sheet[log_dict[key_str]+row].value # 表格实际内容
        for any_tuple in list_dict:
            option,score =any_tuple
            if isinstance(option,tuple): # 如果键是区间（元组）
                if v>=option[0] and v<option[1]:
                    risk_score += score
                    risk_point_c['风险预评估'].append(score)
            else:
                if v == option:
                    risk_score += score 
                    risk_point_c['风险预评估'].append(score)

        replacements['风险预评估'].append(risk_score)
    s_sigma_value = sum(replacements['风险预评估'][:8])
    c_sigma_value = sum(replacements['风险预评估'][8:])
    r_value = s_sigma_value*c_sigma_value
    if r_value<3600:
        r_class='低风险'
    elif r_value>=3600 and r_value<7800:
        r_class='中风险'
    elif r_value>=7800 and r_value<12600:
        r_class='较高风险'
    else:
        r_class='高风险'

    replacements['文本']+=[
        ('+预评估失效可能性得分',s_sigma_value),
        ('+预评估失效后果得分',c_sigma_value),
        ('+预评估风险值',r_value),
        ('+预评估风险等级',r_class)
        ]
    
    replacements['风险预评估打分表—失效可能性']=[]
    temp_list=[
        ('管道名称',f'管道名称：{global_name}'),('记录编号',f'记录编号：{report_name}'),
        # ('+评估日期',sheet[log_dict['评估日期']+row].value),
        ('&100',s_sigma_value)
        ]
    for lst,point in enumerate(risk_point_s['风险预评估']):
        temp_list+=[(f'&{lst+1}',point)]
    replacements['风险预评估打分表—失效可能性']+=temp_list
    
    replacements['风险预评估打分表—失效后果']=[]
    temp_list=[
        ('管道名称',f'管道名称：{global_name}'),('记录编号',f'记录编号：{report_name}'),
        # ('+评估日期',sheet[log_dict['评估日期']+row].value),
        ('&100',c_sigma_value)
        ]
    for lst,point in enumerate(risk_point_c['风险预评估']):
        temp_list += [(f'&{lst+1}',point)]
    replacements['风险预评估打分表—失效后果']+=temp_list


    replacements['风险再评估']=[]
    row = rows[1]
    for key_str,list_dict in RISKY_EVA_S.items():
        risk_score:int=0 
        for son_key,son_tuple in list_dict.items():
            v= sheet[log_dict[son_key]+row].value # 表格实际内容
            for option,score in son_tuple:
                if isinstance(option,tuple): # 如果键是区间（元组）
                    if v>=option[0] and v<option[1]:
                        risk_point_s['风险再评估'].append(score)
                        risk_score += score
                else:
                    if v == option:
                        risk_point_s['风险再评估'].append(score)
                        risk_score += score
        replacements['风险再评估'].append(risk_score)
    for key_str,list_dict in RISKY_EVA_C.items():
        risk_score:int=0 
        v= sheet[log_dict[key_str]+row].value # 表格实际内容
        for any_tuple in list_dict:
            option,score = any_tuple
            if isinstance(option,tuple): # 如果键是区间（元组）
                if v>=option[0] and v<option[1]:
                    risk_score = score
                    risk_point_s['风险再评估'].append(score)
            else:
                if v == option:
                    risk_score = score 
                    risk_point_c['风险再评估'].append(score)

        replacements['风险再评估'].append(risk_score)
    s_sigma_value = sum(replacements['风险再评估'][:8])
    c_sigma_value = sum(replacements['风险再评估'][8:])
    r_value = s_sigma_value*c_sigma_value
    if r_value<3600:
        r_class='低风险'
    elif r_value>=3600 and r_value<7800:
        r_class='中风险'
    elif r_value>=7800 and r_value<12600:
        r_class='较高风险'
    else:
        r_class='高风险'
    replacements['文本']+=[
        ('+再评估失效可能性得分',s_sigma_value),
        ('+再评估失效后果得分',c_sigma_value),
        ('+再评估风险值',r_value),
        ('+再评估风险等级',r_class),
        ('+再评估风险等级',r_class),
        ('+再评估风险等级',r_class)
        ]
    replacements['风险再评估打分表—失效可能性']=[]
    temp_list=[
        ('管道名称',f'管道名称：{global_name}'),('记录编号',f'记录编号：{report_name}'),
        # ('+评估日期',sheet[log_dict['评估日期']+row].value),
        ('&100',s_sigma_value)
        ]
    for lst,point in enumerate(risk_point_s['风险预评估']):
        temp_list+=[(f'&{lst+1}',point)]
    replacements['风险再评估打分表—失效可能性']+=temp_list
    
    replacements['风险再评估打分表—失效后果']=[]
    temp_list=[
        ('管道名称',f'管道名称：{global_name}'),('记录编号',f'记录编号：{report_name}'),
        # ('+评估日期',sheet[log_dict['评估日期']+row].value),
        ('&100',c_sigma_value)
        ]
    for lst,point in enumerate(risk_point_c['风险再评估']):
        temp_list += [(f'&{lst+1}',point)]
    replacements['风险再评估打分表—失效后果']+=temp_list
    return replacements

def make_all_replacement_index(workbook,report_name,gd_dict):
    """管道基本信息：报告编号、管道名称、管道长度等"""
    replacements:list = []
    sheet = workbook['管段清单']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    e_n =len(rows)
    years_set:set[int]={sheet[log_dict['实际使用年限']+row].value for row in rows if sheet[log_dict['实际使用年限']+row].value !='不明' and sheet[log_dict['实际使用年限']+row].value is not None}
    sheet = workbook['管道基本信息']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    row = rows[0]
    # lenth:int = sheet[log_dict['大于20年长度']+row].value + sheet[log_dict['小于20年长度']+row].value
    lenth:int = sheet[log_dict['管道长度']+row].value
    replacements += [
                ('+记录编号',report_name),
                ('+使用单位',sheet[log_dict['使用单位']+row].value),
                ('+管道名称',sheet[log_dict['管道名称']+row].value),
                ('+单位地址',sheet[log_dict['单位地址']+row].value),
                ('+安全管理人员',sheet[log_dict['安全管理人员']+row].value),
                ('+联系电话',sheet[log_dict['联系电话']+row].value),
                # ('+邮政编码',sheet[log_dict['邮政编码']+row].value),
                ('+邮政编码',610000),
                ('+压力管道代码',sheet[log_dict['压力管道代码']+row].value),
                ('+工程总数',e_n),
                ('+管段总数',len(gd_dict)),
                ('+投运年限',f'{min(years_set)}—{max(years_set)}年' ),
                ('+管道长度',lenth/1000),
                 ]
    # CONFIG['审核人'] = sheet[log_dict['审核人']+row].value

    # sheet = workbook['管段清单']
    # log_dict:dict =rg.get_col_in_sheet(sheet)
    # rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # all_set:set[str]=set()
    # years_list:list[float]=[]
    # for row in rows:
    #     year = sheet[log_dict['实际使用年限']+row].value
    #     if year !='不明' and year is not None:
    #         years_list.append(year)
    #     v_str:str = sheet[log_dict['管道规格']+row].value
    #     s=v_str.split(',')
    #     for key in s:
    #         all_set.add(key)
    # replacements += [
    #     # ('+整体管道规格',','.join(map(str,all_set))),
    #     ('+工程总数',len(rows)),
    #     # ('+投运年限',f'{min(years_list)}—{max(years_list)}年'),
    # ]
    
    sheet = workbook['风险评估']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    if CONFIG['是否检验签字']:
        replacements += [
            ('+预评估日期',sheet[log_dict['评估日期']+rows[0]].value),
            ('+资料审查日期',sheet[log_dict['评估日期']+rows[0]].value),
            ('+最早日期',sheet[log_dict['评估日期']+rows[0]].value),
            ('+再评估日期',sheet[log_dict['评估日期']+rows[1]].value),
            ('+完成日期',sheet[log_dict['评估日期']+rows[1]].value),
        ]
    else:
        replacements += [
            ('+预评估日期','年  月  日'),
            ('+资料审查日期','年  月  日'),
            ('+最早日期','年  月  日'),
            ('+再评估日期','年  月  日'),
            ('+完成日期','年  月  日'),
            ('+检验日期','年  月  日')
        ]
    if CONFIG['是否审核签字']:
        replacements += [('+审核日期',CONFIG['审核日期'])]
    else:
        replacements += [('+审核日期','年  月  日')]
    return replacements

def do_replace_in_son_report(doc,any_dict):
    """执行分项报告表格写入"""
    i:int = 0
    j:int = 0
    k:int = 0
    l:int = 1
    m:int = 1
    n:int = 0
    w:int = 0
    for table in doc.Tables:
        title_name:str = table.Title
    
        if title_name == '资料审查记录':
            rg.replace_text_in_table(doc,table,any_dict['资料审查记录'][n],'资料审查记录生成')  
            n+=1
        elif title_name == '宏观检查记录':
            rg.replace_text_in_table(doc,table,any_dict['宏观检查记录'][i],'宏观检查记录生成')  
            i+=1
        elif title_name == '开挖检验记录':
            rg.replace_text_in_table(doc,table,any_dict['开挖检验记录'][j],'开挖检测记录生成')  
            j+=1
        elif title_name == '穿、跨越检查记录' and any_dict['穿、跨越检查记录']:
            rg.replace_text_in_table(doc,table,any_dict['穿、跨越检查记录'][k],'穿、跨越检查记录生成')  
            k+=1
        elif title_name == '开挖记录首页':
            rg.replace_text_in_table(doc,table,any_dict['开挖记录首页'][w],'开挖检验记录生成')  
            w+=1
        elif title_name == '风险预评估记录':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险预评估'][:8]:
                cell = son_table.Cell(2,l)
                cell.Range.Text = score
                l+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险预评估'][8:]:
                cell = son_table.Cell(2,l-8)
                cell.Range.Text = score
                l+=1
        elif title_name=='风险预评估打分表—失效可能性':
            rg.replace_text_in_table(doc,table,any_dict['风险预评估打分表—失效可能性'],'风险评估打分表—失效可能性')
        elif title_name=='风险预评估打分表—失效后果':
            rg.replace_text_in_table(doc,table,any_dict['风险预评估打分表—失效后果'],'风险评估打分表—失效后果')
        elif title_name == '风险再评估记录':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险再评估'][:8]:
                cell = son_table.Cell(2,m)
                cell.Range.Text = score
                m+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险再评估'][8:]:
                cell = son_table.Cell(2,m-8)
                cell.Range.Text = score
                m+=1
        elif title_name=='风险再评估打分表—失效可能性':
            rg.replace_text_in_table(doc,table,any_dict['风险再评估打分表—失效可能性'],'风险评估打分表—失效可能性')
        elif title_name=='风险再评估打分表—失效后果':
            rg.replace_text_in_table(doc,table,any_dict['风险再评估打分表—失效后果'],'风险评估打分表—失效后果')
        else:
            pass

def do_replace_all_pic(doc,pic_dict:dict,path:str):
    """执行所有图片的替换"""
    i:int = 0
    j:int = 0
    k:int = 0
    for shape in doc.InlineShapes:
        tag:str = shape.Title 
        if tag == '开挖检验照片':
            for ex_name in ['.jpg','.png','.jpeg']:
                # f_path:str = f"{path}\\管网840\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\新繁\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\大丰\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\郫三司\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                f_path:str = f"{CONFIG['数据源所在']}\\开挖照片\\{pic_dict['开挖检验照片'][j]}{ex_name}"
                if os.path.exists(f_path):
                    rg.replace_pictue(doc,f_path,shape,120)
                    break
            j+=1
        else:
            if CONFIG['是否检验签字']:
                if tag == '签字':
                    if pic_dict['签字'][i]=='空白':
                        pass
                    else:
                        rg.replace_pictue(doc,f"{CONFIG['签名图片所在']}\\{pic_dict['签字'][i]}.png",shape)
                    i+=1
            if CONFIG['是否审核签字']:
                if tag == '审核签字':
                    rg.replace_pictue(doc,f"{CONFIG['签名图片所在']}\\{CONFIG['审核人']}.jpg",shape)
        
#   检测数据整理
def sort_out_data(workbook:Workbook,report_name:str)->dict[str,str]:
    """按照“管道组织关系”中的内容，返回此报告的（管道编码——名称元组）字典"""
    f_dict:dict[str,str]={}
    sheet = workbook['管段清单']
    log_dict= rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    for row in rows:
        f_dict[str(sheet[log_dict['管道编码']+row].value)]=sheet[log_dict['工程名称']+row].value
    return f_dict
    
def solo_main(report_name:str,workbook:Workbook,word,path:str):
    replacements_dict:dict = {}
    replacements_list:list[tuple] = []
    doc_modle_path = f"{CONFIG['模板文件']}"
    try:
        doc = word.Documents.Open(doc_modle_path)
        gd_dict=sort_out_data(workbook,report_name)
        print('生成替换用文本')
        replacements_dict |= make_replacement_index(workbook,report_name,gd_dict)
        replacements_list += make_all_replacement_index(workbook,report_name,gd_dict) 
       
        print('替换内容')
        do_replace( doc , replacements_dict['文本'],replacements_list )
        
        print('扩张分项报告表格')
        times_dict = expand_all_tables(workbook, doc, report_name,replacements_dict)

        print('填写分项报告表格')
        do_replace_in_son_report(doc,replacements_dict)
    
        print('编制图片替换索引') 
        sign_dict=make_sign_dig_log(workbook,doc,report_name,path,replacements_dict,times_dict,gd_dict)

        print('替换所有图片')
        do_replace_all_pic(doc,sign_dict,path)

        # 移动到文档的末端
        selection = word.Selection
        selection.EndKey(6)  # 6 表示 wdStory，即整个文档

        # 更新文档中的所有域
        doc.Fields.Update()
        
        output_file = f"{CONFIG['输出文件所在']}\\{report_name}原始记录.docx"
        doc.SaveAs2(output_file, FileFormat=16)  # 16 表示docx 17 表示 PDF
        # output_file = f"{config['输出文件']}\\{report_name}原始记录.pdf"
        # doc.SaveAs2(output_file, FileFormat=17)  
        print(f"文档已保存为：{output_file}")

    except Exception as ex:
        traceback.print_exc()
        if doc is not None:
            doc.SaveAs2(f"{CONFIG['输出文件所在']}\\error_{report_name}.docx",FileFormat =16)
            print(f"{report_name}发生错误！")
            doc.Saved =True
            doc.Close(SaveChanges=False)
            raise ex
    finally:
        if doc is not None:
            doc.Close(SaveChanges=False)

if __name__ == '__main__':
    CONFIG:dict[str,str]={
        '模板文件':'E:\\BaiduSyncdisk\\成渝特检\\模板文件与生成程序\\记录、报告生成\\PE管\\PE管原始记录模板.docx',
        '数据源所在':'E:\\BaiduSyncdisk\\成渝特检\\模板文件与生成程序\\记录、报告生成\\PE管\\管网840',
        '签名图片所在':'E:\\BaiduSyncdisk\\成渝特检\\模板文件与生成程序\\记录、报告生成\\PE管\\电子签名',
        '输出文件':'E:\\BaiduSyncdisk\\成渝特检\\模板文件与生成程序\\记录、报告生成\\PE管\\输出文件',
    }
    set_list:list[tuple[int,str,str]]=[
        (2,'模板文件','docx',r'E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\PE管\空港\PE管原始记录模板——空港.docx'),
        (0,'数据源所在','',r'E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\PE管\空港'),
        (0,'签名图片所在','',r'E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\PE管\电子签名'),
        (0,'输出文件所在','',r'E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\PE管\记录文件\空港'),
        (3,'是否检验签字',False,True), 
        (3,'是否审核签字',False,True), 
        (4,'审核人','','付飞'), 
        (4,'审核日期','','2025年12月19日'), 
    ]
    CONFIG = interraction_terminal.set_argumments(set_list)
    app_type = rg.check_office_installation()
    if app_type == None:
        print('未找到合适的应用以打开文档')

    path:str = os.getcwd()
    
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\犀浦\\犀浦_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\新繁\\新繁_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\大丰\\大丰_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\管网840\\管网840_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\郫三司\\郫三司_原始数据.xlsx" )
    workbook:Workbook = openpyxl.load_workbook(f"{CONFIG['数据源所在']}\\原始数据.xlsx" )

    print('读取模板文件')
    
    if app_type == "office":
        word = win32.Dispatch("Word.Application")
    elif app_type == "wps":
        word = win32.Dispatch("Kwps.Application")
    word.Visible = False  # 不显示 Word 窗口，加快处理速度
    word.DisplayAlerts = 0  # 关闭警告信息
    # 全局关闭拼写/语法检查
    word.Options.CheckSpellingAsYouType = False   # 关闭实时拼写检查
    word.Options.CheckGrammarAsYouType = False    # 关闭实时语法检查
    word.Options.ContextualSpeller = False        # 关闭上下文拼写检查（Word 2010+）
    #   初始化完成
    sheet=workbook['管道基本信息']
    log_dict =rg.get_col_in_sheet(sheet)
    all_names:set[str]=set(cell.value for cell in sheet[log_dict['报告编号']] if (cell.value and cell.value.startswith('DGB')))
    
    # for report_name in ['DGB2025111CD',]:
    for report_name in sorted(list(all_names))[:]:
        try:
            solo_main(report_name,workbook,word,path)
        except Exception as e:
            print('有错误发生')
    word.Quit()
        # finally:
        #     continue
    # report_name = 'DGB2025001CD'
    # solo_main(report_name,workbook,word,path)


